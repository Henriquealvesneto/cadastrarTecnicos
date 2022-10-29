import openpyxl
from tkinter import *
from tkinter import ttk
from tkinter import messagebox

    #Criando a planilha que vai receber os dados (wb)
wb = openpyxl.Workbook()
wb.create_sheet('Técnicos Cadastrados')
sheet = wb['Técnicos Cadastrados']

def excel():
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40

    sheet.cell(row=1, column=1).value = "CPF"
    sheet.cell(row=1, column=2).value = "Nome Completo"
    sheet.cell(row=1, column=3).value = "Telefone"
    sheet.cell(row=1, column=4).value = "Turno"
    sheet.cell(row=1, column=5).value = "Equipe"
    sheet.cell(row=1, column=6).value = "Email"

def focus1(evento):
    cpf.focus_set()


def focus2(evento):
    nome_completo.focus_set()


def focus3(evento):
    telefone.focus_set()


def focus4(evento):
    turno.focus_set()


def focus5(evento):
    equipe.focus_set()


def focus6(evento):
    email.focus_set()


def clear():
    cpf.delete(0, END)
    nome_completo.delete(0, END)
    telefone.delete(0, END)
    turno.delete(0, END)
    equipe.delete(0, END)
    email.delete(0, END)


def inserir():
    if (cpf.get() == "" or cpf.get() == " " and
            nome_completo.get() == "" or nome_completo.get() == " " and
            telefone.get() == "" or telefone.get() == " " and
            turno.get() == "" or turno.get() == " " or turno.get() and
            equipe.get() == "" or equipe.get() == " " and
            email.get() == "" or email.get() == " "):
        messagebox.showerror("Error", "Voce deve preencher todos os campos")
    elif (turno.get() != 'Manhã' and turno.get() != "Tarde" and turno.get() != "Noite"):
        messagebox.showerror("Error", "Este Turno não existe, favor preencher corretamente")
    else:
        resposta = messagebox.askquestion("Tem certeza?", "Cadastrar técnico " + nome_completo.get() + "?")
        if resposta == "yes":
            messagebox.showinfo("Sucesso", "Técnico cadastrado com sucesso!")
            wb = openpyxl.load_workbook('testetecnico1.xlsx')
            sheet = wb['Técnicos Cadastrados']
            linha_atual = sheet.max_row
            coluna_atual = sheet.max_column

            sheet.cell(row=linha_atual + 1, column=1).value = cpf.get()
            sheet.cell(row=linha_atual + 1, column=2).value = nome_completo.get()
            sheet.cell(row=linha_atual + 1, column=3).value = telefone.get()
            sheet.cell(row=linha_atual + 1, column=4).value = turno.get()
            sheet.cell(row=linha_atual + 1, column=5).value = equipe.get()
            sheet.cell(row=linha_atual + 1, column=6).value = email.get()
            wb.save('testetecnico1.xlsx')
            cpf.focus_set()
        else:
            messagebox.showwarning("Cancelado", "Técnico não cadastrado!")
        clear()


if __name__ == "__main__":
    root = Tk()

    root.configure(background='light blue')

    root.title("Cadastro de Técnicos")
    root.geometry("500x300")

    excel()

    titulo = Label(root, text="Cadastro de Técnicos", bg="light blue")

    cpf = Label(root, text="CPF", bg="light blue")

    nome_completo = Label(root, text="Nome Completo", bg="light blue")

    telefone = Label(root, text="Telefone", bg="light blue")

    turno = Label(root, text="Turno", bg="light blue")

    equipe = Label(root, text="Equipe", bg="light blue")

    email = Label(root, text="Email", bg="light blue")

    titulo.grid(row=0, column=1)
    cpf.grid(row=1, column=0)
    nome_completo.grid(row=2, column=0)
    telefone.grid(row=3, column=0)
    turno.grid(row=4, column=0)
    equipe.grid(row=5, column=0)
    email.grid(row=6, column=0)

    cpf = Entry(root)
    nome_completo = Entry(root)
    telefone = Entry(root)
    turno = Entry(root)
    equipe = Entry(root)
    email = Entry(root)

    cpf.bind("<Return>", focus1)

    nome_completo.bind("<Return>", focus2)

    telefone.bind("<Return>", focus3)

    turno.bind("<Return>", focus4)

    equipe.bind("<Return>", focus5)

    email.bind("<Return>", focus6)

    cpf.grid(row=1, column=1, ipadx="100")
    nome_completo.grid(row=2, column=1, ipadx="100")
    telefone.grid(row=3, column=1, ipadx="100")
    turnoEscolha = ["Manhã", "Tarde", "Noite"]
    turno = ttk.Combobox(root, values=turnoEscolha)
    turno.set("Manhã")
    turno.grid(row=4,column=1,ipadx="90")
    equipe.grid(row=5, column=1, ipadx="100")
    email.grid(row=6, column=1, ipadx="100")

    excel()

    cadastrar = Button(root, text="Cadastrar", fg="White",
                    bg="Black", command=inserir)
    cadastrar.grid(row=8, column=1)

    root.mainloop()